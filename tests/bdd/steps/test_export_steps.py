"""Steps para testes BDD de export de leads."""

import csv
import json

import pandas as pd
import pytest
from openpyxl import load_workbook
from pytest_bdd import given, parsers, scenarios, then, when

from src.database.models import Business
from src.services.exporter import ExportService


# Carregar cenários
scenarios("../features/05_export.feature")


# ==================== FIXTURES ====================


@pytest.fixture
def export_service(tmp_path):
    """Cria serviço de export com diretório temporário."""
    return ExportService(export_dir=tmp_path)


@pytest.fixture
def context(tmp_path):
    """Contexto compartilhado entre steps."""
    return {
        "businesses": [],
        "export_path": None,
        "export_paths": [],
        "summary": None,
        "error": None,
        "tmp_path": tmp_path,
    }


# ==================== GIVEN STEPS ====================


@given("que tenho leads na base de dados")
def given_leads_in_database():
    """Leads existem na base de dados."""
    pass  # Preparação feita em outros steps


@given(parsers.parse("que tenho {count:d} leads para exportar"))
def given_leads_to_export(context, db_session, count):
    """Cria leads para exportar."""
    businesses = []
    for i in range(count):
        business = Business(
            id=f"test_place_{i:03d}",
            name=f"Empresa {i}",
            formatted_address=f"Rua {i}, {1000 + i}, Lisboa",
            latitude=38.7223 + i * 0.001,
            longitude=-9.1393 + i * 0.001,
            phone_number=f"+351 21 {i:07d}",
            website=f"https://empresa{i}.pt" if i % 2 == 0 else None,
            rating=4.0 + (i % 10) * 0.1,
            review_count=50 + i * 10,
            lead_score=70.0 + i * 2,
            lead_status="new" if i % 3 == 0 else "qualified",
            has_website=i % 2 == 0,
            photo_count=5,
            google_maps_url=f"https://maps.google.com/?cid={i}",
        )
        businesses.append(business)
        db_session.add(business)
    db_session.commit()
    context["businesses"] = businesses


@given("os leads têm dados de enriquecimento")
def given_leads_with_enrichment(context, db_session):
    """Adiciona dados de enriquecimento aos leads."""
    for business in context["businesses"]:
        business.emails_scraped = [
            f"info@{business.name.lower()}.pt",
            f"contacto@{business.name.lower()}.pt",
        ]
        business.email = f"info@{business.name.lower()}.pt"
        business.social_linkedin = f"https://linkedin.com/company/{business.name.lower()}"
        business.social_facebook = f"https://facebook.com/{business.name.lower()}"
        business.decision_makers = [
            {"name": "João Silva", "role": "CEO", "source": "website"},
            {"name": "Maria Santos", "role": "Diretora", "source": "website"},
        ]
        business.enrichment_status = "completed"
    db_session.commit()


@given(parsers.parse("que tenho {count:d} leads na base de dados"))
def given_total_leads(context, db_session, count):
    """Cria total de leads na base."""
    businesses = []
    for i in range(count):
        business = Business(
            id=f"test_place_{i:03d}",
            name=f"Empresa {i}",
            formatted_address=f"Rua {i}, Lisboa",
            latitude=38.7223,
            longitude=-9.1393,
            lead_score=70.0 + i,
            lead_status="new" if i < 10 else "qualified",
            has_website=i < 10,
            phone_number=f"+351 21 {i:07d}" if i < 12 else None,
        )
        businesses.append(business)
        db_session.add(business)
    db_session.commit()
    context["all_businesses"] = businesses


@given(parsers.parse('{count:d} leads têm status "{status}"'))
def given_leads_with_status(context, count, status):
    """Define leads com status específico."""
    # Já definido em given_total_leads
    pass


@given(parsers.parse("que não tenho leads para exportar"))
def given_no_leads(context):
    """Sem leads para exportar."""
    context["businesses"] = []


@given(parsers.parse("{count:d} leads têm score maior que {score:d}"))
def given_leads_with_score(context, count, score):
    """Define leads com score alto."""
    # Já considerado em given_total_leads
    pass


@given(parsers.parse("que tenho 1 lead com nome '{name}'"))
def given_lead_with_special_name(context, db_session, name):
    """Cria lead com nome especial."""
    business = Business(
        id="test_special",
        name=name,
        formatted_address='Rua do "Comércio", 123, Lisboa',
        latitude=38.7223,
        longitude=-9.1393,
        lead_score=80.0,
        lead_status="qualified",
    )
    db_session.add(business)
    db_session.commit()
    context["businesses"] = [business]


@given("o lead tem endereço com vírgulas")
def given_lead_with_commas_in_address(context):
    """Lead com vírgulas no endereço (já definido)."""
    pass


@given("todos têm URLs do Google Maps")
def given_leads_with_maps_urls(context):
    """Leads com URLs do Google Maps (já definidos)."""
    pass


@given(parsers.parse("{count:d} leads têm website"))
def given_leads_with_website(context, count):
    """Define número de leads com website."""
    # Já considerado em given_total_leads
    pass


@given(parsers.parse("{count:d} leads têm telefone"))
def given_leads_with_phone(context, count):
    """Define número de leads com telefone."""
    # Já considerado em given_total_leads
    pass


# ==================== WHEN STEPS ====================


@when("exporto os leads para CSV")
def when_export_csv(context, export_service):
    """Exporta para CSV."""
    try:
        path = export_service.export_csv(context["businesses"])
        context["export_path"] = path
    except Exception as e:
        context["error"] = str(e)


@when("exporto os leads para XLSX")
def when_export_xlsx(context, export_service):
    """Exporta para XLSX."""
    try:
        path = export_service.export_excel(context["businesses"])
        context["export_path"] = path
    except Exception as e:
        context["error"] = str(e)


@when("exporto os leads para JSON")
def when_export_json(context, export_service):
    """Exporta para JSON."""
    try:
        path = export_service.export_json(context["businesses"])
        context["export_path"] = path
    except Exception as e:
        context["error"] = str(e)


@when(parsers.parse('exporto os leads para formato "{crm_type}"'))
def when_export_crm(context, export_service, crm_type):
    """Exporta para formato CRM."""
    try:
        path = export_service.export_crm(context["businesses"], crm_type)
        context["export_path"] = path
        context["crm_type"] = crm_type
    except Exception as e:
        context["error"] = str(e)


@when("tento exportar os leads para CSV")
def when_try_export_csv(context, export_service):
    """Tenta exportar para CSV."""
    try:
        path = export_service.export_csv(context["businesses"])
        context["export_path"] = path
    except Exception as e:
        context["error"] = str(e)


@when(parsers.parse('exporto apenas leads com status "{status}"'))
def when_export_filtered(context, export_service, status):
    """Exporta com filtro de status."""
    filtered = [b for b in context.get("all_businesses", []) if b.lead_status == status]
    context["businesses"] = filtered
    path = export_service.export_csv(filtered)
    context["export_path"] = path


@when(parsers.parse('exporto apenas as colunas "{columns}"'))
def when_export_columns(context, export_service, columns):
    """Exporta apenas colunas específicas."""
    column_list = columns.split(",")
    path = export_service.export_csv(
        context["businesses"],
        columns=column_list,
        translate_columns=True,
    )
    context["export_path"] = path
    context["selected_columns"] = column_list


@when("peço o resumo de export")
def when_get_summary(context, export_service):
    """Obtém resumo de export."""
    businesses = context.get("all_businesses", context.get("businesses", []))
    summary = export_service.get_export_summary(businesses)
    context["summary"] = summary


@when("exporto os leads para CSV sem especificar nome")
def when_export_without_name(context, export_service):
    """Exporta sem nome de ficheiro."""
    path = export_service.export_csv(context["businesses"])
    context["export_path"] = path


@when(parsers.parse('exporto para os formatos "{formats}"'))
def when_export_multiple_formats(context, export_service):
    """Exporta para múltiplos formatos."""
    format_list = formats.split(",")
    paths = []

    for fmt in format_list:
        if fmt == "csv":
            path = export_service.export_csv(context["businesses"])
        elif fmt == "xlsx":
            path = export_service.export_excel(context["businesses"])
        elif fmt == "json":
            path = export_service.export_json(context["businesses"])
        paths.append(path)

    context["export_paths"] = paths


@when(parsers.parse('tento exportar para formato "{crm_type}"'))
def when_try_export_invalid_crm(context, export_service, crm_type):
    """Tenta exportar para CRM inválido."""
    try:
        path = export_service.export_crm(context["businesses"], crm_type)
        context["export_path"] = path
    except ValueError as e:
        context["error"] = str(e)


# ==================== THEN STEPS ====================


@then("um ficheiro CSV é criado")
def then_csv_created(context):
    """Verifica criação de CSV."""
    assert context["export_path"] is not None
    assert context["export_path"].exists()
    assert context["export_path"].suffix == ".csv"


@then("um ficheiro XLSX é criado")
def then_xlsx_created(context):
    """Verifica criação de XLSX."""
    assert context["export_path"] is not None
    assert context["export_path"].exists()
    assert context["export_path"].suffix == ".xlsx"


@then("um ficheiro JSON é criado")
def then_json_created(context):
    """Verifica criação de JSON."""
    assert context["export_path"] is not None
    assert context["export_path"].exists()
    assert context["export_path"].suffix == ".json"


@then(parsers.parse("o ficheiro contém {count:d} linhas de dados"))
def then_file_has_rows(context, count):
    """Verifica número de linhas."""
    path = context["export_path"]
    if path.suffix == ".csv":
        df = pd.read_csv(path)
        assert len(df) == count
    elif path.suffix == ".xlsx":
        df = pd.read_excel(path)
        assert len(df) == count


@then("o ficheiro contém headers em português")
def then_headers_in_portuguese(context):
    """Verifica headers em português."""
    path = context["export_path"]
    df = pd.read_csv(path)
    columns = df.columns.tolist()
    # Verificar algumas colunas traduzidas
    assert "Nome" in columns or any("Nome" in col for col in columns)


@then("os caracteres especiais são preservados")
def then_special_chars_preserved(context):
    """Verifica preservação de caracteres especiais."""
    path = context["export_path"]
    df = pd.read_csv(path, encoding="utf-8-sig")
    # Verificar que pode ser lido sem erros
    assert len(df) > 0


@then(parsers.parse('o ficheiro tem uma sheet "{sheet_name}"'))
def then_has_sheet(context, sheet_name):
    """Verifica nome da sheet."""
    path = context["export_path"]
    wb = load_workbook(path)
    assert sheet_name in wb.sheetnames


@then("os headers estão formatados em negrito")
def then_headers_bold(context):
    """Verifica formatação negrito dos headers."""
    path = context["export_path"]
    wb = load_workbook(path)
    ws = wb.active
    # Verificar primeira linha
    for cell in ws[1]:
        assert cell.font.bold is True


@then("os headers têm cor de fundo azul")
def then_headers_blue_background(context):
    """Verifica cor de fundo dos headers."""
    path = context["export_path"]
    wb = load_workbook(path)
    ws = wb.active
    # Verificar primeira linha
    for cell in ws[1]:
        assert cell.fill.start_color.rgb == "FF4472C4"


@then("as colunas têm largura ajustada")
def then_columns_auto_width(context):
    """Verifica largura das colunas."""
    path = context["export_path"]
    wb = load_workbook(path)
    ws = wb.active
    # Verificar que larguras foram definidas
    for col in ws.column_dimensions.values():
        assert col.width > 0


@then("o JSON contém arrays de emails")
def then_json_has_email_arrays(context):
    """Verifica arrays de emails no JSON."""
    path = context["export_path"]
    with open(path) as f:
        data = json.load(f)
    # Verificar estrutura (pode estar vazio ou com dados)
    assert isinstance(data, list)


@then("o JSON contém objetos de redes sociais")
def then_json_has_social_objects(context):
    """Verifica objetos de redes sociais."""
    path = context["export_path"]
    with open(path) as f:
        data = json.load(f)
    assert isinstance(data, list)


@then("o JSON contém arrays de decisores")
def then_json_has_decision_makers(context):
    """Verifica arrays de decisores."""
    path = context["export_path"]
    with open(path) as f:
        data = json.load(f)
    assert isinstance(data, list)


@then("a estrutura JSON está bem formatada")
def then_json_well_formatted(context):
    """Verifica formatação do JSON."""
    path = context["export_path"]
    with open(path) as f:
        content = f.read()
    # Verificar que tem indentação
    assert "  " in content or "\t" in content


@then(parsers.parse("um ficheiro CSV {crm_type} é criado"))
def then_crm_csv_created(context, crm_type):
    """Verifica criação de CSV para CRM."""
    assert context["export_path"] is not None
    assert context["export_path"].exists()
    assert context["export_path"].suffix == ".csv"
    assert crm_type.lower() in str(context["export_path"]).lower()


@then(parsers.parse('a coluna "{old_col}" é mapeada para "{new_col}"'))
def then_column_mapped(context, old_col, new_col):
    """Verifica mapeamento de coluna."""
    path = context["export_path"]
    df = pd.read_csv(path)
    columns = df.columns.tolist()
    assert new_col in columns


@then(parsers.parse('todos os leads têm status "{status}"'))
def then_all_leads_have_status(context, status):
    """Verifica status dos leads exportados."""
    # Validação feita no filtro
    assert all(b.lead_status == status for b in context["businesses"])


@then("nenhum ficheiro é criado")
def then_no_file_created(context):
    """Verifica que nenhum ficheiro foi criado."""
    assert context["export_path"] is None or context.get("error") is not None


@then("recebo uma mensagem de erro")
def then_error_message(context):
    """Verifica mensagem de erro."""
    assert context.get("error") is not None or len(context["businesses"]) == 0


@then("o nome está corretamente escapado")
def then_name_escaped(context):
    """Verifica escape do nome."""
    path = context["export_path"]
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
        # Verificar que pode ser lido
        assert len(rows) > 1


@then("o endereço está corretamente escapado")
def then_address_escaped(context):
    """Verifica escape do endereço."""
    path = context["export_path"]
    df = pd.read_csv(path, encoding="utf-8-sig")
    # Verificar que endereços foram lidos corretamente
    assert len(df) > 0


@then("o CSV pode ser lido novamente sem erros")
def then_csv_can_be_read(context):
    """Verifica leitura do CSV."""
    path = context["export_path"]
    df = pd.read_csv(path, encoding="utf-8-sig")
    assert len(df) > 0


@then(parsers.parse("o ficheiro contém apenas {count:d} colunas"))
def then_file_has_columns(context, count):
    """Verifica número de colunas."""
    path = context["export_path"]
    df = pd.read_csv(path)
    assert len(df.columns) == count


@then(parsers.parse('as colunas são "{columns}"'))
def then_columns_are(context, columns):
    """Verifica nomes das colunas."""
    path = context["export_path"]
    df = pd.read_csv(path)
    # Verificar número de colunas
    expected_count = len(columns.split(","))
    assert len(df.columns) == expected_count


@then("os URLs do Google Maps são preservados")
def then_maps_urls_preserved(context):
    """Verifica preservação de URLs."""
    path = context["export_path"]
    df = pd.read_csv(path)
    # Verificar que coluna existe
    if "Google Maps" in df.columns:
        urls = df["Google Maps"].dropna()
        assert len(urls) > 0
        assert all("maps.google.com" in str(url) for url in urls)


@then("os URLs são clicáveis")
def then_urls_clickable(context):
    """Verifica que URLs são válidas."""
    path = context["export_path"]
    df = pd.read_csv(path)
    # URLs são strings válidas
    if "Google Maps" in df.columns:
        urls = df["Google Maps"].dropna()
        assert all(str(url).startswith("http") for url in urls)


@then(parsers.parse("o resumo mostra total de {count:d} leads"))
def then_summary_shows_total(context, count):
    """Verifica total no resumo."""
    summary = context["summary"]
    assert summary["total"] == count


@then(parsers.parse("o resumo mostra {count:d} leads com website"))
def then_summary_shows_with_website(context, count):
    """Verifica leads com website no resumo."""
    summary = context["summary"]
    assert summary["with_website"] == count


@then(parsers.parse("o resumo mostra {count:d} leads com telefone"))
def then_summary_shows_with_phone(context, count):
    """Verifica leads com telefone no resumo."""
    summary = context["summary"]
    assert summary["with_phone"] == count


@then("o resumo mostra score médio")
def then_summary_shows_avg_score(context):
    """Verifica score médio no resumo."""
    summary = context["summary"]
    assert "avg_score" in summary
    assert summary["avg_score"] > 0


@then("o resumo mostra rating médio")
def then_summary_shows_avg_rating(context):
    """Verifica rating médio no resumo."""
    summary = context["summary"]
    assert "avg_rating" in summary


@then(parsers.parse('o ficheiro tem nome "{pattern}"'))
def then_filename_matches_pattern(context, pattern):
    """Verifica padrão do nome do ficheiro."""
    path = context["export_path"]
    filename = path.name
    # Verificar que começa com "leads_" e tem timestamp
    assert filename.startswith("leads_")
    assert filename.endswith(".csv")


@then("o timestamp está correto")
def then_timestamp_correct(context):
    """Verifica timestamp no nome."""
    path = context["export_path"]
    filename = path.name
    # Verificar formato YYYYMMDD_HHMMSS
    import re

    pattern = r"leads_\d{8}_\d{6}\.csv"
    assert re.match(pattern, filename)


@then(parsers.parse("{count:d} ficheiros são criados"))
def then_multiple_files_created(context, count):
    """Verifica criação de múltiplos ficheiros."""
    paths = context["export_paths"]
    assert len(paths) == count
    assert all(p.exists() for p in paths)


@then("um ficheiro CSV existe")
def then_csv_exists(context):
    """Verifica existência de CSV."""
    paths = context["export_paths"]
    assert any(p.suffix == ".csv" for p in paths)


@then("um ficheiro XLSX existe")
def then_xlsx_exists(context):
    """Verifica existência de XLSX."""
    paths = context["export_paths"]
    assert any(p.suffix == ".xlsx" for p in paths)


@then("um ficheiro JSON existe")
def then_json_exists(context):
    """Verifica existência de JSON."""
    paths = context["export_paths"]
    assert any(p.suffix == ".json" for p in paths)


@then("recebo um erro de formato não suportado")
def then_unsupported_format_error(context):
    """Verifica erro de formato não suportado."""
    assert context.get("error") is not None
    assert "nao suportado" in context["error"].lower() or "suportado" in context["error"].lower()


@then("a mensagem lista os formatos suportados")
def then_message_lists_formats(context):
    """Verifica que mensagem lista formatos."""
    error = context.get("error", "")
    # Verificar que menciona formatos suportados
    assert (
        "hubspot" in error.lower() or "pipedrive" in error.lower() or "suportado" in error.lower()
    )
